import cv2
import os
#import time
import logging
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side
from deepface import DeepFace
from deep_sort_realtime.deepsort_tracker import DeepSort
#from ultralytics import YOLO  # Import YOLOv11m

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Paths
db_path = "database"
excel_path = "face_tracking.xlsx"  # Excel file to store tracking data

# Initialize YOLOv11m face detector
#detector = YOLO("yolov11m-face.pt")  # Load YOLO model

# Initialize DeepSORT tracker
tracker = DeepSort(max_age=2, n_init=3, nn_budget=200, embedder="mobilenet")

# Dictionary to store DeepFace identity and tracking time
identity_to_elapsed_time = {}
identity_to_last_seen = {}
identity_active = {}

def format_time(seconds):
    """Convert seconds to HH:MM:SS format"""
    hours = int(seconds // 3600)
    minutes = int((seconds % 3600) // 60)
    secs = int(seconds % 60)
    return f"{hours:02}:{minutes:02}:{secs:02}"

def recognize_face(face):
    """Recognize face using DeepFace"""
    try:
        recognition_results = DeepFace.find(
            img_path=face,
            db_path=db_path,
            model_name="ArcFace",
            enforce_detection=False,
            detector_backend='yolov8',
            align=True,
            threshold=0.65,
            normalization='ArcFace'
        )
        if len(recognition_results[0]) > 0:
            face_rec_conf = recognition_results[0]["distance"].values[0]
            if face_rec_conf < 0.65:
                identity = recognition_results[0]["identity"].values[0].split("/")[-1].split(".")[0]
            else:
                identity = "Unknown"
        else:
            identity = "Unknown"
    except Exception as e:
        logging.error(f"DeepFace recognition error: {e}")
        identity = "Unknown"

    for folder in os.listdir(db_path):
        if folder in identity:
            identity = folder
            break

    return identity

def get_new_sheet_name():
    """Generate a new sheet name based on existing sheets"""
    if not os.path.exists(excel_path):
        return "Run_1"

    with pd.ExcelFile(excel_path) as xls:
        existing_sheets = xls.sheet_names
        run_numbers = [int(sheet.split("_")[1]) for sheet in existing_sheets if sheet.startswith("Run_")]
        new_run_number = max(run_numbers, default=0) + 1
        return f"Run_{new_run_number}"

def save_to_excel():
    """Save identity tracking times to a new sheet in the existing Excel file"""
    df = pd.DataFrame.from_dict(identity_to_elapsed_time, orient="index", columns=["Total Time (seconds)"])
    df["Formatted Time"] = df["Total Time (seconds)"].apply(format_time)

    sheet_name = get_new_sheet_name()

    if os.path.exists(excel_path):
        with pd.ExcelWriter(excel_path, mode="a", engine="openpyxl") as writer:
            df.to_excel(writer, index_label="Identity", sheet_name=sheet_name)
    else:
        df.to_excel(excel_path, index_label="Identity", sheet_name=sheet_name)

    adjust_column_width(excel_path, sheet_name)
    logging.info(f"Tracking data saved to Excel under sheet: {sheet_name}")
    return sheet_name

def adjust_column_width(excel_path, sheet_name):
    """Adjust column width for better readability"""
    try:
        wb = load_workbook(excel_path)
        sheet = wb[sheet_name]

        for col in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            sheet.column_dimensions[col_letter].width = adjusted_width

        wb.save(excel_path)
        logging.info(f"Column width adjusted for sheet: {sheet_name}")

    except Exception as e:
        logging.error(f"Error adjusting column width: {e}")

def format_excel(excel_path, sheet_name):
    """Format the Excel sheet"""
    try:
        wb = load_workbook(excel_path)
        sheet = wb[sheet_name]

        no_border = Border(left=Side(style=None), right=Side(style=None),
                           top=Side(style=None), bottom=Side(style=None))

        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = no_border

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.value = round(cell.value, 2)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = no_border

        wb.save(excel_path)
        print(f"Formatting applied successfully to sheet: {sheet_name}")

    except Exception as e:
        print(f"Error in formatting: {e}")

# Open video
cap = cv2.VideoCapture('video3_hd.mp4')
fps = cap.get(cv2.CAP_PROP_FPS)
frame_count = 0

while cap.isOpened():
    ret, frame = cap.read()
    if not ret:
        break

    frame_count += 1
    current_time = frame_count / fps  # Time based on FPS

    #results = detector(frame)
    detections = []
    detected_identities = set()

    recognition_results = DeepFace.find(
        img_path=frame,
        db_path="database/",
        model_name="ArcFace",
        detector_backend='yolov8',
        enforce_detection=False,
        align=True
    )

    # Get face locations separately (bounding boxes)
    faces = DeepFace.extract_faces(
        img_path=frame,
        detector_backend='yolov8',
        enforce_detection=False
    )

    if isinstance(recognition_results, list):
        for i, face_info in enumerate(faces):
            if i >= len(recognition_results):
                break
            df = recognition_results[i]
            if df.empty:
                identity = "Unknown"
            else:
                identity_path = df.iloc[0]["identity"]
                distance = round(df.iloc[0]["distance"], 3)
                identity = os.path.basename(identity_path).split('.')[0]
                # identity = f"{identity} ({distance})"

            for folder in os.listdir(db_path):
                index = identity.find(folder)
                if index != -1:
                    identity = folder
                    break

            region = face_info["facial_area"]
            x, y, w, h = region["x"], region["y"], region["w"], region["h"]

            detections.append(([x, y, w, h], 0, identity))
            detected_identities.add(identity)

    '''        
    for result in results:
        for box, conf in zip(result.boxes.xyxy, result.boxes.conf):
            x1, y1, x2, y2 = map(int, box)
            padding = 30
            x1 = x1 - padding
            y1 = y1 - padding
            x2 = x2 + padding
            y2 = y2 + padding
            face_img = frame[y1:y2, x1:x2]
            identity = recognize_face(face_img)
            detections.append(([x1, y1, x2 - x1, y2 - y1], float(conf), identity))
            detected_identities.add(identity)
    '''

    tracked_objects = tracker.update_tracks(detections, frame=frame)
    active_identities = set()

    for track in tracked_objects:
        if not track.is_confirmed():
            continue

        identity = track.det_class
        bbox = track.to_ltrb()
        active_identities.add(identity)

        if identity not in identity_to_elapsed_time:
            identity_to_elapsed_time[identity] = 0.0
            identity_to_last_seen[identity] = current_time
            identity_active[identity] = True
        elif identity in detected_identities:
            if not identity_active.get(identity, False):
                identity_to_last_seen[identity] = current_time
                identity_active[identity] = True
            elapsed = current_time - identity_to_last_seen[identity]
            identity_to_elapsed_time[identity] += elapsed
            identity_to_last_seen[identity] = current_time

        elapsed_time_str = format_time(identity_to_elapsed_time[identity])
        color = (0, 255, 0) if identity != "Unknown" else (0, 0, 255)
        x1, y1, x2, y2 = map(int, bbox)
        cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)
        cv2.putText(frame, f"{identity} | {elapsed_time_str}", (x1, y1 - 10),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 2)

    for identity in identity_active:
        if identity not in active_identities:
            identity_active[identity] = False

    cv2.imshow("Tracking", frame)
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# Save and format on exit
sheet_name = save_to_excel()
format_excel(excel_path, sheet_name)
cap.release()
cv2.destroyAllWindows()
